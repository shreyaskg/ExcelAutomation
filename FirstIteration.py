import openpyxl
import datetime
import pandas as pd
from datetime import date
from GetStandardCollegeNamedistance import GetStandardName
from GetStandardCollegeNamedistance import GetDistance
from GetStandardCollegeNamedistance import GetParticipants
from GetAge import GetDays
from GetAge import GetAge

# storing today's date in the variable today
def FirstIteration(file):
    print("First iteration running")
    skeletal_file = 'skeletal_file.xlsx'
    output = 'output.xlsx'
    today = datetime.datetime.now()
    # filename argument should contain the file path if in different directory
    wb = openpyxl.load_workbook(filename = file)
    # opening the file to be written into
    wb_write = openpyxl.load_workbook(filename = skeletal_file)

    # get the reference to the required sheet in the xlsx file
    o_sheet = wb["Callsheet"]

    write_sheet = wb_write["Final Data"]
    # Loading the information in the week1.xlsx file
    data = []
    for i in range(2,o_sheet.max_row + 3):
        individual_data = []
        for column in o_sheet[i]:
            individual_data.append(column.value)
        data.append(individual_data)

    print("Got all the data, setting weeks")

    weeks = 1

    for i in range(1,len(data) + 1):

        if type(data[i-1][15]) == type(today):
            weeks = (today - data[i-1][15]).days
            if weeks % 7 == 0:
                weeks = weeks / 7
            else:
                # integer division
                weeks = weeks // 7 + 1
            # as per the weeks mentioned
            if weeks > 40:
                weeks = 40
            weeks = 17 + weeks*3

            # static data
            print(data[i-1][0])
            write_sheet.cell(i+1,1).value = data[i-1][0]
            write_sheet.cell(i+1,2).value = data[i - 1][1]
            write_sheet.cell(i + 1, 3).value = data[i - 1][3]
            write_sheet.cell(i + 1, 11).value = data[i - 1][20]
            write_sheet.cell(i + 1, 13).value = data[i - 1][22]
            write_sheet.cell(i + 1, 14).value = data[i - 1][23]
            write_sheet.cell(i + 1, 15).value = data[i - 1][24]


            # dynamic data
            write_sheet.cell(i+1,weeks + 1).value = data[i-1][13]
            write_sheet.cell(i+1,weeks + 2).value = data[i-1][17]
            write_sheet.cell(i + 1, weeks).value = data[i - 1][12]

            # College dynamic field
            college = data[i-1][4]
            standard_college = GetStandardName(college)
            write_sheet.cell(i + 1,5).value =  standard_college
            distance = GetDistance(standard_college)
            write_sheet.cell(i+1,17).value = distance

            # Days dynamic field
            day = data[i-1][6]
            ageofregistration = GetDays(day)
            write_sheet.cell(i+1,6).value = ageofregistration

            day = data[i-1][7]
            dateofprogram = -GetDays(day)
            write_sheet.cell(i+1,7).value = dateofprogram

            # Current week response dynamic field
            write_sheet.cell(i+1,8).value = data[i-1][12]
            write_sheet.cell(i+1,9).value = data[i-1][13]
            write_sheet.cell(i+1,10).value = data[i - 1][17]

            # Age Dynamic field
            write_sheet.cell(i+1,12).value = GetAge(data[i-1][21])

            # Date of attendance
            write_sheet.cell(i+1,19).value = data[i-1][7]

    wb_write.save(output)
# FirstIteration('TempWeek.xlsx')


