import openpyxl
import ast

def SetParticipants(file):
    write = openpyxl.load_workbook(filename=file)
    sheet = write['Final Data']
    colleges = {}
    for i in range(2, sheet.max_row + 3):
        if sheet[i][4].value not in colleges and sheet[i][4].value :
            colleges[sheet[i][4].value] = 1
        elif sheet[i][4].value in colleges and sheet[i][4].value not in [None,"undefined","#N/A"] :
            colleges[sheet[i][4].value] += 1
    print(colleges)
    for i in range(2,sheet.max_row-1):
        if sheet[i][4].value not in ['0', 0, None, "#N/A", "undefined"]:
            sheet.cell(i,18).value = colleges[sheet[i][4].value]
        else:
            sheet.cell(i, 18).value = 1
    write.save('output.xlsx')
# SetParticipants('output.xlsx')