import openpyxl
import datetime
import pandas as pd
from datetime import date

# storing today's date in the variable today
def CreatingSkeletalFile(file,weeks):

    today = datetime.datetime.now()
    # filename argument should contain the file path if in different directory
    wb = openpyxl.load_workbook(filename = file, data_only = True)
    # opening the file to be written into
    wb_write = openpyxl.load_workbook(filename = 'Weekly callresponses.xlsx',data_only = True)
    counter = 0
    # get the reference to the required sheet in the xlsx file
    o_sheet = wb.get_sheet_by_name("Callsheet")
    write_sheet = wb_write.get_sheet_by_name("Final Data")
    j = 1
    # creating columns (30) containing call response | call status | calls
    weeks = weeks*3 + 20
    for i in range(20,weeks,3):
        write_sheet.cell(1,i).value = 'Week ' + str(j) + ' call response'
        write_sheet.cell(1,i+1).value = 'Week ' + str(j) + ' call status'
        write_sheet.cell(1,i+2).value = 'Week ' + str(j) + ' calls'
        j += 1

    wb_write.save("skeletal_file.xlsx")
# CreatingSkeletalFile('Week 1.xlsx',30)