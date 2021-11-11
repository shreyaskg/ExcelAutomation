import openpyxl
import datetime
import pandas as pd
from datetime import date
from GetStandardCollegeNamedistance import GetStandardName
from GetStandardCollegeNamedistance import GetDistance
from GetStandardCollegeNamedistance import GetParticipants
from GetAge import GetDays
from GetAge import GetAge

# Loading the already written week 1 file
def SecondIterationOnwards(file):
    output = 'output.xlsx'
    wb_write = openpyxl.load_workbook(filename = output)
    write_sheet = wb_write.get_sheet_by_name("Final Data")

    # storing today's date in the variable today
    today = datetime.datetime.now()
    wb_2 = openpyxl.load_workbook(filename = file)
    o_sheet_2 = wb_2["Callsheet"]
    data2 = []
    phone_numbers = []
    for i in range(2,o_sheet_2.max_row-3):
        individual_data = []
        for column in o_sheet_2[i]:
            individual_data.append(column.value)
        data2.append(individual_data)
        phone_numbers.append(individual_data[1])

    #print(data2[0][1])
    new_row_index = []
    for i in range(1,len(data2)):
        new_row_index.append(i)

    wb = openpyxl.load_workbook(filename = 'Week 1.xlsx')
    o_sheet = wb["Callsheet"]
    data = []
    for i in range(2,o_sheet.max_row + 3):
        individual_data = []
        for column in o_sheet[i]:
            individual_data.append(column.value)
        data.append(individual_data)

    phone_number_duplicate = []
    counter = 0
    flag = 0
    for i in range(1,len(data2)):
        phone_number_new = data2[i][1]
        for j in range(1,len(data)):
            phone_number_existing = data[j][1]
            if phone_number_new == phone_number_existing:
                phone_number_duplicate.append(phone_number_existing)
                counter += 1
                # print(f"{phone_number_new} : {phone_number_existing}")
                # calculate the week difference and set in the row_number
                row_number = j
                if data2 is None:
                    continue
                try:
                    if type(data2[i - 1][15]) == type(today):
                        weeks = (today - data2[i - 1][15]).days

                        if weeks % 7 == 0:
                            weeks = weeks / 7
                        else:
                            # integer division
                            weeks = weeks // 7 + 1
                        # as per the weeks mentioned
                        if weeks > 30:
                            weeks = 30

                        weeks = 17 + weeks * 3
                        write_sheet.cell(j+2, weeks).value = (data2[i][12])
                        write_sheet.cell(j+2, weeks + 1).value = (data2[i][13])
                        write_sheet.cell(j+2, weeks + 2).value = (data2[i][17])

                        # Update the latest responses

                        write_sheet.cell(j+2, 8).value = (data2[i][12])
                        write_sheet.cell(j+2, 9).value = (data2[i][13])
                        write_sheet.cell(j+2, 10).value = (data2[i][17])

                        # Dynamic fields which are also static, as we will be retaining the latest response
                        write_sheet.cell(j+2, 19).value = data[i - 1][7]

                        day = data[i][6]
                        ageofregistration = GetDays(day)
                        write_sheet.cell(j+2, 6).value = ageofregistration

                        day = data[i][7]
                        dateofprogram = -GetDays(day)
                        write_sheet.cell(j+2, 7).value = dateofprogram


                        flag = 1
                except:
                    pass
    total_rows = o_sheet.max_row
    print(total_rows)
    flag = 0
    for i in range(len(phone_numbers)):
        for j in range(len(phone_number_duplicate)):
            if phone_numbers[i] not in phone_number_duplicate:
                flag = 1
        if flag == 1:
            if data2[i][15] is None:
                continue
            if type(data2[i][15]) == type(today):
                weeks = (today - data2[i][15]).days
                if weeks % 7 == 0:
                    weeks = weeks / 7
                else:
                    # integer division
                    weeks = weeks // 7 + 1
                # as per the weeks mentioned
                if weeks > 30:
                    weeks = 30
                weeks = 15 + weeks * 3
                # static fields
                write_sheet.cell(total_rows - 1, 1).value = data[i][0]
                write_sheet.cell(total_rows - 1, 2).value = data[i][1]
                write_sheet.cell(total_rows - 1, 3).value = data[i][3]
                write_sheet.cell(total_rows - 1, 11).value = data[i][20]
                write_sheet.cell(total_rows - 1, 13).value = data[i][22]
                write_sheet.cell(total_rows - 1, 14).value = data[i][23]
                write_sheet.cell(total_rows - 1, 15).value = data[i][24]

                # dynamic data
                write_sheet.cell(total_rows - 1, weeks + 1).value = data[i][13]
                write_sheet.cell(total_rows - 1, weeks + 2).value = data[i][17]
                write_sheet.cell(total_rows - 1, weeks).value = data[i][12]

                # College dynamic field
                # college = data[i][4]
                # standard_college = GetStandardName(college)
                # write_sheet.cell(total_rows - 1, 5).value = standard_college
                # distance = GetDistance(standard_college)
                # write_sheet.cell(total_rows - 1, 17).value = distance

                # Days dynamic field
                day = data[i][6]
                ageofregistration = GetDays(day)
                write_sheet.cell(total_rows - 1, 6).value = ageofregistration

                day = data[i][7]
                dateofprogram = -GetDays(day)
                write_sheet.cell(total_rows - 1, 7).value = dateofprogram

                # Current week response dynamic field
                write_sheet.cell(total_rows - 1, 8).value = data[i][12]
                write_sheet.cell(total_rows - 1, 9).value = data[i][13]
                write_sheet.cell(total_rows - 1, 10).value = data[i][17]

                # Age Dynamic field
                write_sheet.cell(total_rows - 1, 12).value = GetAge(data[i][21])

                # Date of attendance
                write_sheet.cell(total_rows - 1, 19).value = data[i][7]
                total_rows += 1
        flag = 0

    wb_write.save('output2.xlsx')
SecondIterationOnwards('Week 2.xlsx')