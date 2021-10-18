import openpyxl
import datetime
import pandas as pd
from datetime import date

# Loading the already written week 1 file
def SecondIterationOnwards(file):
    output = 'output.xlsx'
    wb_write = openpyxl.load_workbook(filename = output)
    write_sheet = wb_write.get_sheet_by_name("Final Data")

    # storing today's date in the variable today
    today = datetime.datetime.now()
    wb_2 = openpyxl.load_workbook(filename = file)
    o_sheet_2 = wb_2["Callsheet"]
    data2 = []
    phone_numbers = []
    for i in range(2,o_sheet_2.max_row + 3):
        individual_data = []
        for column in o_sheet_2[i]:
            individual_data.append(column.value)
        data2.append(individual_data)
        phone_numbers.append(individual_data[1])

    #print(data2[0][1])
    new_row_index = []
    for i in range(1,len(data2)):
        new_row_index.append(i)

    wb = openpyxl.load_workbook(filename = 'Week 1.xlsx')
    o_sheet = wb["Callsheet"]
    data = []
    for i in range(2,o_sheet.max_row + 3):
        individual_data = []
        for column in o_sheet[i]:
            individual_data.append(column.value)
        data.append(individual_data)

    phone_number_duplicate = []
    counter = 0
    flag = 0
    for i in range(1,len(data2)):
        phone_number_new = data2[i][1]
        for j in range(1,len(data)):
            phone_number_existing = data[j][1]
            if phone_number_new == phone_number_existing:
                phone_number_duplicate.append(phone_number_existing)
                counter += 1
                # print(f"{phone_number_new} : {phone_number_existing}")
                # calculate the week difference and set in the row_number
                row_number = j
                if data2 is None:
                    continue
                try:
                    if str(type(data2[i - 1][15])) == "<class 'datetime.datetime'>" and len(str(type(data[i - 1][15]))) > 5:
                        weeks = (today - data2[i - 1][15]).days

                        if weeks % 7 == 0:
                            weeks = weeks / 7
                        else:
                            # integer division
                            weeks = weeks // 7 + 1
                        # as per the weeks mentioned
                        if weeks > 30:
                            weeks = 30

                        weeks = 15 + weeks * 3
                        write_sheet.cell(j+2, weeks).value = (data2[i][12])
                        write_sheet.cell(j+2, weeks + 1).value = (data2[i][13])
                        write_sheet.cell(j+2, weeks + 2).value = (data2[i][17])
                        flag = 1
                except:
                    pass
    total_rows = o_sheet.max_row
    flag = 0
    for i in range(len(phone_numbers)):
        for j in range(len(phone_number_duplicate)):
            if phone_numbers[i] not in phone_number_duplicate:
                flag = 1
        if flag == 1:
            if data2[i][15] is None:
                continue
            if str(type(data2[i][15])) == "<class 'datetime.datetime'>" and len(str(type(data2[i][15]))) > 5:
                weeks = (today - data2[i][15]).days
                if weeks % 7 == 0:
                    weeks = weeks / 7
                else:
                    # integer division
                    weeks = weeks // 7 + 1
                # as per the weeks mentioned
                if weeks > 30:
                    weeks = 30
                weeks = 15 + weeks * 3
                # static fields
                write_sheet.cell(total_rows -1 ,1).value = data2[i][0]
                write_sheet.cell(total_rows - 1, 2).value = data2[i][1]
                write_sheet.cell(total_rows - 1, 3).value = data2[i][2]
                write_sheet.cell(total_rows - 1, 5).value = data2[i][4]
                write_sheet.cell(total_rows - 1, 9).value = data2[i][20]
                write_sheet.cell(total_rows - 1, 11).value = data2[i][22]
                write_sheet.cell(total_rows - 1, 12).value = data2[i][23]
                write_sheet.cell(total_rows - 1, 13).value = data2[i][24]
                # dynamic fields
                write_sheet.cell(total_rows - 1,weeks).value = data2[i][12]
                write_sheet.cell(total_rows - 1, weeks + 1).value = data2[i][13]
                write_sheet.cell(total_rows - 1, weeks + 2).value = data2[i][17]
                total_rows += 1
        flag = 0

    wb_write.save(output)